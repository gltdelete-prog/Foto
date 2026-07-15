#!/usr/bin/env python3
"""Generate August 2026 driver schedule Excel for SERDECHNOV TEAM."""

from calendar import monthrange
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

YEAR = 2026
MONTH = 8
RATE_PER_TRIP = 4000
NIGHT_BONUS = 500
NORM_SHK = 1341
THRESHOLD_SHK = 1300

DRIVERS = {
    "ivan": "Иван",
    "alexey": "Алексей",
    "v3": "Водитель 3",
}

# 5/2: weekday 0=Mon .. 6=Sun
OFF_DAYS = {
    "ivan": {5, 6},      # Сб, Вс
    "alexey": {0, 1},    # Пн, Вт
    "v3": {2, 3},        # Ср, Чт
}

DAY_NAMES = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
MONTH_NAMES = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]

ROTATION = [
    {"day": "ivan", "night": "alexey"},
    {"day": "alexey", "night": "v3"},
    {"day": "v3", "night": "ivan"},
]

TRIPS = {
    "night": {"time": "02:00", "truck": "Машина 2", "count": 1, "label": "Ночной"},
    "day_morning": {"time": "06:00", "truck": "Машина 1", "count": 0.5, "label": "Утренний"},
    "day_evening": {"time": "18:00", "truck": "Машина 1", "count": 0.5, "label": "Вечерний"},
}

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E79")
sub_font = Font(size=10, color="666666")
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

fill_work = PatternFill("solid", fgColor="E2EFDA")
fill_off = PatternFill("solid", fgColor="FCE4D6")
fill_night = PatternFill("solid", fgColor="DDEBF7")
fill_day = PatternFill("solid", fgColor="FFF2CC")
fill_reserve = PatternFill("solid", fgColor="E7E6E6")
fill_weekend = PatternFill("solid", fgColor="F2F2F2")


def is_working(driver_key: str, weekday: int) -> bool:
    return weekday not in OFF_DAYS[driver_key]


def week_index(d: date) -> int:
    return (d.day - 1) // 7


def pick_driver(preferred: str, working: list[str], exclude: set[str]) -> str | None:
    if preferred in working and preferred not in exclude:
        return preferred
    for key in ["ivan", "alexey", "v3"]:
        if key in working and key not in exclude:
            return key
    return None


def build_day_schedule(d: date) -> dict:
    wd = d.weekday()
    working = [k for k in DRIVERS if is_working(k, wd)]
    rot = ROTATION[week_index(d) % 3]

    day_driver = pick_driver(rot["day"], working, set())
    night_driver = pick_driver(rot["night"], working, {day_driver} if day_driver else set())

    if day_driver is None and working:
        day_driver = working[0]
    if night_driver is None:
        night_driver = pick_driver(
            rot["night"],
            [k for k in working if k != day_driver],
            set(),
        )

    reserve = [k for k in working if k not in {day_driver, night_driver}]

    return {
        "date": d,
        "weekday": wd,
        "working": working,
        "day_driver": day_driver,
        "night_driver": night_driver,
        "reserve": reserve,
    }


def style_range(ws, cell_range, fill=None, font=None, alignment=None, border_=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border_:
                cell.border = border_


def build_workbook() -> Workbook:
    wb = Workbook()
    days_in_month = monthrange(YEAR, MONTH)[1]
    all_days = [date(YEAR, MONTH, day) for day in range(1, days_in_month + 1)]
    schedules = [build_day_schedule(d) for d in all_days]

    # --- Sheet 1: Calendar ---
    ws = wb.active
    ws.title = "График август"
    ws.sheet_view.showGridLines = True

    ws.merge_cells("A1:J1")
    ws["A1"] = f"SERDECHNOV TEAM — график рейсов на {MONTH_NAMES[MONTH]} {YEAR}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"3 рейса/сутки · 2 машины · маршрут 7 · норма {NORM_SHK} ШК · "
        f"вызов машины при ≥{THRESHOLD_SHK} ШК · {RATE_PER_TRIP} ₽/рейс"
    )
    ws["A2"].font = sub_font
    ws["A2"].alignment = center

    headers = [
        "Дата", "День", "02:00\n(Маш.2)", "06:00\n(Маш.1)", "18:00\n(Маш.1)",
        "Резерв", "Рейсов\nвсего", "ШК\n(план)", "Выручка\n₽", "Примечание",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    stats = {k: {"trips": 0, "night_trips": 0, "work_days": 0, "off_days": 0} for k in DRIVERS}

    for i, sch in enumerate(schedules):
        row = 5 + i
        d = sch["date"]
        wd = sch["weekday"]

        night_name = DRIVERS[sch["night_driver"]] if sch["night_driver"] else "—"
        day_name = DRIVERS[sch["day_driver"]] if sch["day_driver"] else "—"
        reserve = ", ".join(DRIVERS[k] for k in sch["reserve"]) or "—"

        for key in DRIVERS:
            if key in sch["working"]:
                stats[key]["work_days"] += 1
            else:
                stats[key]["off_days"] += 1

        if sch["night_driver"]:
            stats[sch["night_driver"]]["trips"] += 1
            stats[sch["night_driver"]]["night_trips"] += 1
        if sch["day_driver"]:
            stats[sch["day_driver"]]["trips"] += 2

        note_parts = []
        if len(sch["working"]) == 3:
            note_parts.append("3 водителя — резерв на ПВЗ/техосмотр")
        if wd >= 5:
            note_parts.append("выходные")

        values = [
            d.strftime("%d.%m.%Y"),
            DAY_NAMES[wd],
            night_name,
            day_name,
            day_name,
            reserve,
            3,
            NORM_SHK * 3,
            3 * RATE_PER_TRIP,
            "; ".join(note_parts) if note_parts else "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center if col != 10 else left

        ws.cell(row=row, column=3).fill = fill_night
        ws.cell(row=row, column=4).fill = fill_day
        ws.cell(row=row, column=5).fill = fill_day
        if wd >= 5:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = fill_weekend

    total_row = 5 + len(schedules)
    ws.cell(row=total_row, column=1, value="ИТОГО").font = bold
    ws.cell(row=total_row, column=7, value=len(schedules) * 3).font = bold
    ws.cell(row=total_row, column=8, value=NORM_SHK * 3 * len(schedules)).font = bold
    ws.cell(row=total_row, column=9, value=len(schedules) * 3 * RATE_PER_TRIP).font = bold
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).fill = fill_reserve

    widths = [12, 6, 14, 14, 14, 14, 8, 10, 12, 28]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[4].height = 36

    # Legend
    leg = total_row + 2
    ws.cell(row=leg, column=1, value="Легенда:").font = bold
    legends = [
        (fill_night, "02:00 — ночной рейс (Машина 2)"),
        (fill_day, "06:00 / 18:00 — дневные рейсы (Машина 1, один водитель)"),
        (fill_weekend, "Суббота / воскресенье"),
    ]
    for j, (f, text) in enumerate(legends):
        r = leg + 1 + j
        ws.cell(row=r, column=1).fill = f
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=text).alignment = left

    # --- Sheet 2: Driver summary ---
    ws2 = wb.create_sheet("Зарплата и рейсы")
    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"Сводка по водителям — {MONTH_NAMES[MONTH]} {YEAR}"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center

    h2 = ["Водитель", "Рабочих дней", "Выходных", "Рейсов", "Ночных рейсов", "Зарплата ₽"]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, key in enumerate(["ivan", "alexey", "v3"], 4):
        s = stats[key]
        salary = s["trips"] * RATE_PER_TRIP + s["night_trips"] * NIGHT_BONUS
        row_data = [
            DRIVERS[key],
            s["work_days"],
            s["off_days"],
            s["trips"],
            s["night_trips"],
            salary,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = center

    tr = 7
    ws2.cell(row=tr, column=1, value="ИТОГО").font = bold
    ws2.cell(row=tr, column=2, value=sum(s["work_days"] for s in stats.values())).font = bold
    ws2.cell(row=tr, column=4, value=sum(s["trips"] for s in stats.values())).font = bold
    ws2.cell(row=tr, column=6, value=sum(
        s["trips"] * RATE_PER_TRIP + s["night_trips"] * NIGHT_BONUS for s in stats.values()
    )).font = bold

    ws2.merge_cells("A9:F9")
    ws2["A9"] = (
        f"Формула: рейс × {RATE_PER_TRIP} ₽ + ночная доплата × {NIGHT_BONUS} ₽. "
        "Роли (дневной/ночной) чередуются каждую неделю."
    )
    ws2["A9"].font = sub_font
    ws2["A9"].alignment = left

    for idx, w in enumerate([16, 14, 12, 10, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # --- Sheet 3: 5/2 pattern ---
    ws3 = wb.create_sheet("График 5-2")
    ws3.merge_cells("A1:D1")
    ws3["A1"] = "Постоянный график выходных (5/2)"
    ws3["A1"].font = title_font

    ws3["A3"] = "Водитель"
    ws3["B3"] = "Выходные"
    ws3["C3"] = "Рабочие дни"
    for c in ["A3", "B3", "C3"]:
        ws3[c].font = header_font
        ws3[c].fill = header_fill
        ws3[c].alignment = center
        ws3[c].border = border

    off_info = [
        ("Иван", "Суббота, Воскресенье", "Понедельник – Пятница"),
        ("Алексей", "Понедельник, Вторник", "Среда – Воскресенье"),
        ("Водитель 3", "Среда, Четверг", "Пятница – Вторник"),
    ]
    for i, (name, off, work) in enumerate(off_info, 4):
        ws3.cell(row=i, column=1, value=name).border = border
        ws3.cell(row=i, column=2, value=off).border = border
        ws3.cell(row=i, column=2).fill = fill_off
        ws3.cell(row=i, column=3, value=work).border = border
        ws3.cell(row=i, column=3).fill = fill_work

    ws3.merge_cells("A8:D8")
    ws3["A8"] = "Ротация ролей по неделям"
    ws3["A8"].font = bold

    rot_headers = ["Неделя", "02:00 (ночь)", "06:00 + 18:00 (день)"]
    for col, h in enumerate(rot_headers, 1):
        c = ws3.cell(row=9, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border

    weeks = [
        ("1 (3–9 авг)", "Алексей", "Иван"),
        ("2 (10–16 авг)", "Водитель 3", "Алексей"),
        ("3 (17–23 авг)", "Иван", "Водитель 3"),
        ("4 (24–31 авг)", "Алексей", "Иван"),
    ]
    for i, row in enumerate(weeks, 10):
        for col, val in enumerate(row, 1):
            ws3.cell(row=i, column=col, value=val).border = border

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 28

    # --- Sheet 4: Daily checklist ---
    ws4 = wb.create_sheet("Чек-лист ПВЗ")
    ws4.merge_cells("A1:E1")
    ws4["A1"] = "Ежедневный чек-лист — маршрут 7"
    ws4["A1"].font = title_font

    checklist = [
        ("Время", "Действие", "Порог ШК", "Машина", "Кто проверяет"),
        ("01:00", "Проверить накопление ШК на парковке 7", f"≥ {THRESHOLD_SHK}", "—", "Дежурный ПВЗ"),
        ("01:30", "Выезд на ночную загрузку", f"≥ {THRESHOLD_SHK}", "Машина 2", "Ночной водитель"),
        ("02:00", "Загрузка маршрут 7", str(NORM_SHK), "Машина 2", "Ночной водитель"),
        ("05:30", "Проверить ШК перед утренней загрузкой", f"≥ {THRESHOLD_SHK}", "—", "Дежурный ПВЗ"),
        ("05:30", "Выезд на утреннюю загрузку", f"≥ {THRESHOLD_SHK}", "Машина 1", "Дневной водитель"),
        ("06:00", "Загрузка маршрут 7", str(NORM_SHK), "Машина 1", "Дневной водитель"),
        ("17:30", "Проверить ШК перед вечерней загрузкой", f"≥ {THRESHOLD_SHK}", "—", "Дежурный ПВЗ"),
        ("17:30", "Выезд на вечернюю загрузку", f"≥ {THRESHOLD_SHK}", "Машина 1", "Дневной водитель"),
        ("18:00", "Загрузка маршрут 7", str(NORM_SHK), "Машина 1", "Дневной водитель"),
    ]
    for r, row in enumerate(checklist, 3):
        for c, val in enumerate(row, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = left if c > 1 else center
            if r == 3:
                cell.font = header_font
                cell.fill = header_fill

    for idx, w in enumerate([10, 42, 12, 12, 18], 1):
        ws4.column_dimensions[get_column_letter(idx)].width = w

    return wb


if __name__ == "__main__":
    output = "/workspace/SERDECHNOV-TEAM-grafik-avgust-2026.xlsx"
    build_workbook().save(output)
    print(f"Saved: {output}")
