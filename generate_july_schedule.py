#!/usr/bin/env python3
"""Generate driver schedule Excel for a custom date range."""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

START = date(2026, 7, 20)
END = date(2026, 7, 31)
RATE_PER_TRIP = 4000
NIGHT_BONUS = 500
NORM_SHK = 1341
THRESHOLD_SHK = 1300

DRIVERS = {
    "d1": "1 Водитель",
    "d2": "2 Водитель",
    "d3": "3 Водитель",
}

OFF_DAYS = {
    "d1": {5, 6},   # Сб, Вс
    "d2": {0, 1},   # Пн, Вт
    "d3": {2, 3},   # Ср, Чт
}

DAY_NAMES = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

ROTATION = [
    {"day": "d1", "night": "d2"},
    {"day": "d2", "night": "d3"},
    {"day": "d3", "night": "d1"},
]

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E79")
sub_font = Font(size=10, color="666666")
bold = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
fill_night = PatternFill("solid", fgColor="DDEBF7")
fill_day = PatternFill("solid", fgColor="FFF2CC")
fill_off = PatternFill("solid", fgColor="FCE4D6")
fill_work = PatternFill("solid", fgColor="E2EFDA")
fill_weekend = PatternFill("solid", fgColor="F2F2F2")
fill_total = PatternFill("solid", fgColor="E7E6E6")


def is_working(driver_key: str, weekday: int) -> bool:
    return weekday not in OFF_DAYS[driver_key]


def week_index(d: date) -> int:
    # Rotate every 4 days for fair split on short periods
    return (d - START).days // 4


def pick_driver(preferred: str, working: list[str], exclude: set[str]) -> str | None:
    if preferred in working and preferred not in exclude:
        return preferred
    for key in ["d1", "d2", "d3"]:
        if key in working and key not in exclude:
            return key
    return None


def date_range(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def build_day_schedule(d: date) -> dict:
    wd = d.weekday()
    working = [k for k in DRIVERS if is_working(k, wd)]
    rot = ROTATION[week_index(d) % 3]

    day_driver = pick_driver(rot["day"], working, set())
    night_driver = pick_driver(rot["night"], working, {day_driver} if day_driver else set())

    if day_driver is None and working:
        day_driver = working[0]
    if night_driver is None:
        night_driver = pick_driver(
            rot["night"],
            [k for k in working if k != day_driver],
            set(),
        )

    reserve = [k for k in working if k not in {day_driver, night_driver}]
    return {
        "date": d,
        "weekday": wd,
        "working": working,
        "day_driver": day_driver,
        "night_driver": night_driver,
        "reserve": reserve,
    }


def build_workbook() -> Workbook:
    wb = Workbook()
    all_days = date_range(START, END)
    schedules = [build_day_schedule(d) for d in all_days]
    stats = {k: {"trips": 0, "night_trips": 0, "work_days": 0, "off_days": 0} for k in DRIVERS}

    ws = wb.active
    ws.title = "График 20-31 июля"
    ws.merge_cells("A1:J1")
    ws["A1"] = f"SERDECHNOV TEAM — график {START.strftime('%d.%m.%Y')} – {END.strftime('%d.%m.%Y')}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"3 рейса/сутки · 2 машины · {RATE_PER_TRIP} ₽/рейс + {NIGHT_BONUS} ₽ ночная смена · "
        f"норма {NORM_SHK} ШК"
    )
    ws["A2"].font = sub_font
    ws["A2"].alignment = center

    headers = [
        "Дата", "День", "02:00\n(Маш.2)", "06:00\n(Маш.1)", "18:00\n(Маш.1)",
        "Резерв", "Рейсов", "ШК (план)", "Выручка ₽", "Примечание",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, sch in enumerate(schedules):
        row = 5 + i
        d = sch["date"]
        wd = sch["weekday"]

        night_name = DRIVERS[sch["night_driver"]] if sch["night_driver"] else "—"
        day_name = DRIVERS[sch["day_driver"]] if sch["day_driver"] else "—"
        reserve = ", ".join(DRIVERS[k] for k in sch["reserve"]) or "—"

        for key in DRIVERS:
            if key in sch["working"]:
                stats[key]["work_days"] += 1
            else:
                stats[key]["off_days"] += 1

        if sch["night_driver"]:
            stats[sch["night_driver"]]["trips"] += 1
            stats[sch["night_driver"]]["night_trips"] += 1
        if sch["day_driver"]:
            stats[sch["day_driver"]]["trips"] += 2

        note = "3 водителя — резерв" if len(sch["working"]) == 3 else ""
        if wd >= 5:
            note = (note + "; выходные").strip("; ")

        values = [
            d.strftime("%d.%m.%Y"), DAY_NAMES[wd], night_name, day_name, day_name,
            reserve, 3, NORM_SHK * 3, 3 * RATE_PER_TRIP, note,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = center if col != 10 else left

        ws.cell(row=row, column=3).fill = fill_night
        ws.cell(row=row, column=4).fill = fill_day
        ws.cell(row=row, column=5).fill = fill_day
        if wd >= 5:
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = fill_weekend

    total_row = 5 + len(schedules)
    ws.cell(row=total_row, column=1, value="ИТОГО").font = bold
    ws.cell(row=total_row, column=7, value=len(schedules) * 3).font = bold
    ws.cell(row=total_row, column=8, value=NORM_SHK * 3 * len(schedules)).font = bold
    ws.cell(row=total_row, column=9, value=len(schedules) * 3 * RATE_PER_TRIP).font = bold
    for col in range(1, 11):
        ws.cell(row=total_row, column=col).border = border
        ws.cell(row=total_row, column=col).fill = fill_total

    widths = [12, 6, 14, 14, 14, 14, 8, 10, 12, 24]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Salary sheet
    ws2 = wb.create_sheet("Зарплата")
    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"Зарплата водителей {START.strftime('%d.%m')} – {END.strftime('%d.%m.%Y')}"
    ws2["A1"].font = title_font
    ws2["A1"].alignment = center

    h2 = [
        "Водитель", "Рабочих дней", "Выходных", "Рейсов", "Ночных",
        f"Рейсы ({RATE_PER_TRIP} ₽)", f"Ночь ({NIGHT_BONUS} ₽)", "ИТОГО ₽",
    ]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for i, key in enumerate(["d1", "d2", "d3"], 4):
        s = stats[key]
        trip_pay = s["trips"] * RATE_PER_TRIP
        night_pay = s["night_trips"] * NIGHT_BONUS
        total = trip_pay + night_pay
        row_data = [
            DRIVERS[key], s["work_days"], s["off_days"], s["trips"], s["night_trips"],
            trip_pay, night_pay, total,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = border
            cell.alignment = center

    tr = 7
    ws2.cell(row=tr, column=1, value="ИТОГО").font = bold
    ws2.cell(row=tr, column=4, value=sum(s["trips"] for s in stats.values())).font = bold
    ws2.cell(row=tr, column=8, value=sum(
        s["trips"] * RATE_PER_TRIP + s["night_trips"] * NIGHT_BONUS for s in stats.values()
    )).font = bold

    ws2.merge_cells("A9:H9")
    ws2["A9"] = f"Формула: рейс × {RATE_PER_TRIP} ₽ + ночная смена × {NIGHT_BONUS} ₽"
    ws2["A9"].font = sub_font

    for idx, w in enumerate([14, 12, 10, 8, 10, 14, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # 5/2 sheet
    ws3 = wb.create_sheet("График 5-2")
    ws3["A1"] = "Постоянный график выходных (5/2)"
    ws3["A1"].font = title_font
    rows = [
        ("1 Водитель", "Суббота, Воскресенье", "Понедельник – Пятница"),
        ("2 Водитель", "Понедельник, Вторник", "Среда – Воскресенье"),
        ("3 Водитель", "Среда, Четверг", "Пятница – Вторник"),
    ]
    for col, h in enumerate(["Водитель", "Выходные", "Рабочие дни"], 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
    for i, (a, b, c_val) in enumerate(rows, 4):
        ws3.cell(row=i, column=1, value=a).border = border
        ws3.cell(row=i, column=2, value=b).border = border
        ws3.cell(row=i, column=2).fill = fill_off
        ws3.cell(row=i, column=3, value=c_val).border = border
        ws3.cell(row=i, column=3).fill = fill_work

    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 28

    # Per-driver sheets
    driver_daily: dict[str, list] = {k: [] for k in DRIVERS}
    for sch in schedules:
        d = sch["date"]
        for key in DRIVERS:
            if key not in sch["working"]:
                driver_daily[key].append({
                    "date": d, "status": "ВЫХОДНОЙ", "shifts": [], "pay": 0,
                })
                continue
            shifts = []
            pay = 0
            if sch["night_driver"] == key:
                shifts.append("02:00 — ночной рейс (Маш.2)")
                pay += RATE_PER_TRIP + NIGHT_BONUS
            if sch["day_driver"] == key:
                shifts.append("06:00 — утренний рейс (Маш.1)")
                shifts.append("18:00 — вечерний рейс (Маш.1)")
                pay += RATE_PER_TRIP * 2
            if key in sch["reserve"]:
                shifts.append("Резерв / ПВЗ")
            driver_daily[key].append({
                "date": d, "status": "РАБОТА", "shifts": shifts, "pay": pay,
            })

    for key in ["d1", "d2", "d3"]:
        name = DRIVERS[key]
        ws_d = wb.create_sheet(name[:31])
        ws_d.merge_cells("A1:E1")
        ws_d["A1"] = f"{name} — личный график"
        ws_d["A1"].font = title_font
        ws_d["A1"].alignment = center

        dh = ["Дата", "День", "Смена", "Рейсов", "Заработок ₽"]
        for col, h in enumerate(dh, 1):
            c = ws_d.cell(row=3, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        total_trips = 0
        total_pay = 0
        for i, day in enumerate(driver_daily[key], 4):
            d = day["date"]
            trips = sum(1 for s in day["shifts"] if "рейс" in s)
            shift_text = "\n".join(day["shifts"]) if day["shifts"] else day["status"]
            row_vals = [
                d.strftime("%d.%m.%Y"), DAY_NAMES[d.weekday()],
                shift_text, trips if day["status"] == "РАБОТА" else "—", day["pay"] or "—",
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws_d.cell(row=i, column=col, value=val)
                cell.border = border
                cell.alignment = center if col != 3 else left
            if day["status"] == "ВЫХОДНОЙ":
                for col in range(1, 6):
                    ws_d.cell(row=i, column=col).fill = fill_off
            elif trips >= 2:
                ws_d.cell(row=i, column=3).fill = fill_day
            elif "ночной" in shift_text:
                ws_d.cell(row=i, column=3).fill = fill_night
            total_trips += trips
            total_pay += day["pay"]

        tr = 4 + len(driver_daily[key])
        ws_d.cell(row=tr, column=1, value="ИТОГО").font = bold
        ws_d.cell(row=tr, column=4, value=total_trips).font = bold
        ws_d.cell(row=tr, column=5, value=total_pay).font = bold
        for col in range(1, 6):
            ws_d.cell(row=tr, column=col).border = border
            ws_d.cell(row=tr, column=col).fill = fill_total

        for idx, w in enumerate([12, 6, 36, 8, 14], 1):
            ws_d.column_dimensions[get_column_letter(idx)].width = w

    return wb


if __name__ == "__main__":
    out = "/workspace/SERDECHNOV-grafik-20-31-iyulya-2026.xlsx"
    build_workbook().save(out)
    print(f"Saved: {out}")

    # Print summary
    from datetime import date as d
    days = date_range(START, END)
    print(f"Days: {len(days)}")
    for day in days:
        print(day.strftime("%d.%m.%Y"), DAY_NAMES[day.weekday()])
